import openpyxl
import csv

wb = openpyxl.load_workbook(filename = 'transpose_url.xlsx')
ws = wb.get_sheet_by_name(name = 'Sheet1')

column_count = ws.max_column

gen_list = []

for col in ws.iter_cols(min_col = 2):
    for cell in col:
        gen_list.append(cell.value)

url_list = list(set(gen_list))
num_values = len(url_list)


f = open('new_url_list.csv', 'w')
writer = csv.writer(f, delimiter = ',', quotechar = '"', quoting = csv.QUOTE_MINIMAL)
for url in url_list:
    if url is not None:
        writer.writerow([url])

f.close()
#wb.save('new_data.xlsx')
